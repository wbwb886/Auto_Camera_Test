import os
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from config.device_config import REPORT_PATH

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


class SimpleReport:
    _instance = None  
    _report_file = None  

    def __new__(cls, report_dir=REPORT_PATH):
        """单例模式，保证 pytest 一次运行只有一个 Excel 报告"""
        if cls._instance is None:
            cls._instance = super(SimpleReport, cls).__new__(cls)

            os.makedirs(report_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            cls._report_file = os.path.join(report_dir, f"test_report_{timestamp}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "TestReport"
            headers = ["Case_Name", "Loops", "Result", "Comments"]
            ws.append(headers)

            # 表头样式：黄色背景、加粗、居中、加边框
            header_font = Font(bold=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for i, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=i)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                ws.column_dimensions[get_column_letter(i)].width = 30

            wb.save(cls._report_file)
            logger.info(f"[Report Init] Excel 报告生成: {cls._report_file}")

        return cls._instance

    @property
    def report_file(self):
        return self._report_file

    def add_result(self, case_name, loops, result, comments=""):
        """写入一条测试结果到 Excel"""
        if isinstance(comments, (list, tuple)):
            comments = "\n".join(str(c) for c in comments)
        else:
            comments = comments.replace(";", "\n")

        wb = load_workbook(self._report_file)
        ws = wb.active

        ws.append([case_name, loops, result, comments])

        result_cell = ws.cell(row=ws.max_row, column=3)
        comments_cell = ws.cell(row=ws.max_row, column=4)

        # 换行
        comments_cell.alignment = Alignment(wrap_text=True)

        # 格子边框
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # 整行加边框（四列）
        for col in range(1, 5):
            ws.cell(row=ws.max_row, column=col).border = border

        # 结果背景色
        if result.upper() == "PASS":
            result_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif result.upper() == "FAIL":
            result_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # comments 有内容 → 标红
        if comments.strip():
            comments_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        wb.save(self._report_file)

        logger.info(f"Report updated: {case_name} | {loops} | {result} | {comments}")
